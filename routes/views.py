from django.contrib.auth.decorators import login_required
from django.shortcuts import render, redirect, get_object_or_404
from django.db.models import Count, Avg, Q
from django.contrib import messages
from django.utils import timezone
from django.db.models.functions import TruncDate
from django.http import HttpResponse, JsonResponse
from django.template.loader import get_template
from django.conf import settings
from django.core.mail import send_mail
from django.views.decorators.csrf import csrf_exempt

from .forms import RouteAssignmentForm, ReportFilterForm
from .services import create_optimized_route, mark_stop_collected, mark_stop_skipped, log_route_activity
from .models import Route, RouteStop, RouteActivity, DriverLocation

from alerts.telegram_service import send_telegram_message
from routes.services_reporting import get_report_summary, build_summary_text
from bins.models import Bin, BinReading
from alerts.models import Alert
from account.models import CustomUser

from xhtml2pdf import pisa
from openpyxl.styles import Font
from datetime import timedelta
import io
import openpyxl
import json

@login_required
def route_list_view(request):
    if request.user.role not in ['ADMIN', 'MANAGER']:
        return redirect('role_redirect')
    routes = Route.objects.prefetch_related('stops').order_by('-created_at')
    return render(request, 'routes/route_list.html', {'routes': routes})

@login_required
def generate_route_view(request):
    if request.user.role not in ['ADMIN', 'MANAGER']:
        return redirect('role_redirect')
    if request.method == 'POST':
        route, message = create_optimized_route(created_by=request.user)
        return render(request, 'routes/generate_result.html', {
            'route': route,
            'message': message,
        })
    return render(request, 'routes/generate_route.html')

@login_required
def assign_route_view(request, route_id):
    if request.user.role not in ['ADMIN', 'MANAGER']:
        return redirect('role_redirect')
    route = get_object_or_404(Route, id=route_id)
    if request.method == 'POST':
        form = RouteAssignmentForm(request.POST, instance=route)
        if form.is_valid():
            assigned_route = form.save(commit=False)
            if assigned_route.assigned_driver:
                assigned_route.status = 'ASSIGNED'
            assigned_route.save()
            return redirect('route_list')
    else:
        form = RouteAssignmentForm(instance=route)
    return render(request, 'routes/assign_route.html', {
        'route': route,
        'form': form,
    })

@login_required
def assigned_routes_view(request):
    if request.user.role != 'DRIVER':
        return redirect('role_redirect')
    routes = Route.objects.filter(assigned_driver=request.user).prefetch_related('stops').order_by('-created_at')
    return render(request, 'routes/assigned_routes.html', {'routes': routes})

@login_required
def route_detail_view(request, route_id):
    route = get_object_or_404(Route.objects.prefetch_related('stops__bin'), id=route_id)
    if request.user.role == 'DRIVER':
        if route.assigned_driver != request.user:
            return redirect('role_redirect')
    elif request.user.role not in ['ADMIN', 'MANAGER']:
        return redirect('role_redirect')
    return render(request, 'routes/route_detail.html', {'route': route})

@login_required
def mark_stop_collected_view(request, stop_id):
    stop = get_object_or_404(RouteStop, id=stop_id)
    if request.user.role != 'DRIVER':
        return redirect('role_redirect')
    if stop.route.assigned_driver != request.user:
        return redirect('role_redirect')
    if request.method == 'POST':
        reset_bin = request.POST.get('reset_bin') == 'on'
        progress = mark_stop_collected(stop, user=request.user, reset_bin=reset_bin)
        messages.success(
            request,
            f"Stop marked as collected. Route status: {progress['route_status']}"
        )
        return redirect('route_detail', route_id=stop.route.id)
    return render(request, 'routes/confirm_collect.html', {'stop': stop})

@login_required
def mark_stop_skipped_view(request, stop_id):
    stop = get_object_or_404(RouteStop, id=stop_id)
    if request.user.role != 'DRIVER':
        return redirect('role_redirect')
    if stop.route.assigned_driver != request.user:
        return redirect('role_redirect')
    if request.method == 'POST':
        reason = request.POST.get('reason', '').strip()
        progress = mark_stop_skipped(stop, reason=reason)
        messages.warning(
            request,
            f"Stop marked as skipped. Route status: {progress['route_status']}"
        )
        return redirect('route_detail', route_id=stop.route.id)
    return render(request, 'routes/confirm_skip.html', {'stop': stop})

@login_required
def analytics_dashboard_view(request):
    if request.user.role not in ['ADMIN', 'MANAGER']:
        return redirect('role_redirect')
    total_collections = RouteStop.objects.filter(status='COLLECTED').count()
    completed_routes = Route.objects.filter(status='COMPLETED').count()
    skipped_stops = RouteStop.objects.filter(status='SKIPPED').count()
    average_fill_level = BinReading.objects.aggregate(
        avg_fill=Avg('fill_percentage')
    )['avg_fill'] or 0
    active_alerts = Alert.objects.filter(status='ACTIVE').count()
    offline_bins = Bin.objects.filter(status='OFFLINE').count()
    problematic_bins = Bin.objects.annotate(
        skipped_count=Count('route_stops', filter=Q(route_stops__status='SKIPPED')),
        alert_count=Count('alerts'),
        offline_alert_count=Count(
            'alerts',
            filter=Q(alerts__alert_type='BIN_OFFLINE')
        )
    ).order_by('-skipped_count', '-alert_count', '-offline_alert_count')[:10]
    driver_performance = CustomUser.objects.filter(role='DRIVER').annotate(
        assigned_routes_count=Count('assigned_routes', distinct=True),
        completed_routes_count=Count(
            'assigned_routes',
            filter=Q(assigned_routes__status='COMPLETED'),
            distinct=True
        ),
        collected_stops_count=Count(
            'assigned_routes__stops',
            filter=Q(assigned_routes__stops__status='COLLECTED')
        ),
        skipped_stops_count=Count(
            'assigned_routes__stops',
            filter=Q(assigned_routes__stops__status='SKIPPED')
        ),
    )
    recent_activities = RouteActivity.objects.select_related(
        'route', 'user', 'stop'
    ).order_by('-created_at')[:15]
    start_date = timezone.now().date() - timedelta(days=6)
    alert_trends_qs = (
        Alert.objects.filter(created_at__date__gte=start_date)
        .annotate(day=TruncDate('created_at'))
        .values('day', 'alert_type')
        .annotate(total=Count('id'))
        .order_by('day')
    )
    route_trends_qs = (
        Route.objects.filter(created_at__date__gte=start_date)
        .annotate(day=TruncDate('created_at'))
        .values('day', 'status')
        .annotate(total=Count('id'))
        .order_by('day')
    )
    fill_trends_qs = (
        BinReading.objects.filter(created_at__date__gte=start_date)
        .annotate(day=TruncDate('created_at'))
        .values('day')
        .annotate(avg_fill=Avg('fill_percentage'))
        .order_by('day')
    )
    day_labels = [(start_date + timedelta(days=i)).strftime('%Y-%m-%d') for i in range(7)]
    full_alert_map = {day: 0 for day in day_labels}
    offline_alert_map = {day: 0 for day in day_labels}
    completed_route_map = {day: 0 for day in day_labels}
    in_progress_route_map = {day: 0 for day in day_labels}
    fill_level_map = {day: 0 for day in day_labels}
    for row in alert_trends_qs:
        day = row['day'].strftime('%Y-%m-%d')
        if row['alert_type'] == 'BIN_FULL':
            full_alert_map[day] = row['total']
        elif row['alert_type'] == 'BIN_OFFLINE':
            offline_alert_map[day] = row['total']
    for row in route_trends_qs:
        day = row['day'].strftime('%Y-%m-%d')
        if row['status'] == 'COMPLETED':
            completed_route_map[day] = row['total']
        elif row['status'] == 'IN_PROGRESS':
            in_progress_route_map[day] = row['total']
    for row in fill_trends_qs:
        day = row['day'].strftime('%Y-%m-%d')
        fill_level_map[day] = round(row['avg_fill'] or 0, 2)
    driver_labels = [driver.username for driver in driver_performance]
    driver_collected_data = [driver.collected_stops_count for driver in driver_performance]
    driver_skipped_data = [driver.skipped_stops_count for driver in driver_performance]
    context = {
        'total_collections': total_collections,
        'completed_routes': completed_routes,
        'skipped_stops': skipped_stops,
        'average_fill_level': round(average_fill_level, 2),
        'active_alerts': active_alerts,
        'offline_bins': offline_bins,
        'problematic_bins': problematic_bins,
        'driver_performance': driver_performance,
        'recent_activities': recent_activities,
        'chart_labels': day_labels,
        'full_alert_data': [full_alert_map[day] for day in day_labels],
        'offline_alert_data': [offline_alert_map[day] for day in day_labels],
        'completed_route_data': [completed_route_map[day] for day in day_labels],
        'in_progress_route_data': [in_progress_route_map[day] for day in day_labels],
        'fill_level_data': [fill_level_map[day] for day in day_labels],
        'driver_labels': driver_labels,
        'driver_collected_data': driver_collected_data,
        'driver_skipped_data': driver_skipped_data,
    }
    return render(request, 'routes/analytics_dashboard.html', context)

def get_filtered_report_data(request):
    form = ReportFilterForm(request.GET or None)
    routes = Route.objects.select_related('assigned_driver', 'created_by').prefetch_related('stops__bin').all()
    stops = RouteStop.objects.select_related('route', 'bin', 'route__assigned_driver').all()
    if form.is_valid():
        start_date = form.cleaned_data.get('start_date')
        end_date = form.cleaned_data.get('end_date')
        driver = form.cleaned_data.get('driver')
        bin_obj = form.cleaned_data.get('bin_obj')
        route_status = form.cleaned_data.get('route_status')
        if start_date:
            routes = routes.filter(created_at__date__gte=start_date)
            stops = stops.filter(route__created_at__date__gte=start_date)
        if end_date:
            routes = routes.filter(created_at__date__lte=end_date)
            stops = stops.filter(route__created_at__date__lte=end_date)
        if driver:
            routes = routes.filter(assigned_driver=driver)
            stops = stops.filter(route__assigned_driver=driver)
        if bin_obj:
            routes = routes.filter(stops__bin=bin_obj).distinct()
            stops = stops.filter(bin=bin_obj)
        if route_status:
            routes = routes.filter(status=route_status)
            stops = stops.filter(route__status=route_status)
    summary = {
        'total_routes': routes.count(),
        'completed_routes': routes.filter(status='COMPLETED').count(),
        'in_progress_routes': routes.filter(status='IN_PROGRESS').count(),
        'assigned_routes': routes.filter(status='ASSIGNED').count(),
        'cancelled_routes': routes.filter(status='CANCELLED').count(),
        'total_collections': stops.filter(status='COLLECTED').count(),
        'skipped_stops': stops.filter(status='SKIPPED').count(),
        'pending_stops': stops.filter(status='PENDING').count(),
        'average_fill_level': round(
            BinReading.objects.aggregate(avg_fill=Avg('fill_percentage'))['avg_fill'] or 0, 2
        ),
    }
    problematic_bins = Bin.objects.annotate(
        skipped_count=Count('route_stops', filter=Q(route_stops__status='SKIPPED')),
        alert_count=Count('alerts')
    ).order_by('-skipped_count', '-alert_count')[:10]
    driver_summary = CustomUser.objects.filter(role='DRIVER').annotate(
        assigned_routes_count=Count('assigned_routes', distinct=True),
        completed_routes_count=Count(
            'assigned_routes',
            filter=Q(assigned_routes__status='COMPLETED'),
            distinct=True
        ),
        collected_stops_count=Count(
            'assigned_routes__stops',
            filter=Q(assigned_routes__stops__status='COLLECTED')
        ),
        skipped_stops_count=Count(
            'assigned_routes__stops',
            filter=Q(assigned_routes__stops__status='SKIPPED')
        ),
    )
    return {
        'form': form,
        'routes': routes.order_by('-created_at'),
        'stops': stops.order_by('route__route_name', 'stop_order'),
        'summary': summary,
        'problematic_bins': problematic_bins,
        'driver_summary': driver_summary,
    }

@login_required
def reports_home_view(request):
    if request.user.role not in ['ADMIN', 'MANAGER']:
        return redirect('role_redirect')

    context = get_filtered_report_data(request)
    return render(request, 'routes/reports_home.html', context)


@login_required
def export_reports_pdf_view(request):
    if request.user.role not in ['ADMIN', 'MANAGER']:
        return redirect('role_redirect')

    context = get_filtered_report_data(request)
    template = get_template('routes/reports_pdf.html')
    html = template.render(context)

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="waste_management_report.pdf"'

    pisa_status = pisa.CreatePDF(io.BytesIO(html.encode('utf-8')), dest=response)

    if pisa_status.err:
        return HttpResponse('Error generating PDF report', status=500)

    return response


@login_required
def export_reports_excel_view(request):
    if request.user.role not in ['ADMIN', 'MANAGER']:
        return redirect('role_redirect')

    context = get_filtered_report_data(request)
    routes = context['routes']
    stops = context['stops']
    summary = context['summary']

    workbook = openpyxl.Workbook()

    summary_sheet = workbook.active
    summary_sheet.title = 'Summary'

    summary_sheet['A1'] = 'Waste Management Report Summary'
    summary_sheet['A1'].font = Font(bold=True, size=14)

    row = 3
    for key, value in summary.items():
        summary_sheet[f'A{row}'] = key.replace('_', ' ').title()
        summary_sheet[f'B{row}'] = value
        row += 1

    routes_sheet = workbook.create_sheet(title='Routes')
    headers = ['Route Name', 'Date', 'Status', 'Driver', 'Total Bins', 'Distance (km)', 'Created By']
    for col_num, header in enumerate(headers, 1):
        cell = routes_sheet.cell(row=1, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)

    for row_num, route in enumerate(routes, 2):
        routes_sheet.cell(row=row_num, column=1, value=route.route_name)
        routes_sheet.cell(row=row_num, column=2, value=str(route.date))
        routes_sheet.cell(row=row_num, column=3, value=route.status)
        routes_sheet.cell(row=row_num, column=4, value=str(route.assigned_driver) if route.assigned_driver else '-')
        routes_sheet.cell(row=row_num, column=5, value=route.total_bins)
        routes_sheet.cell(row=row_num, column=6, value=route.total_distance_km)
        routes_sheet.cell(row=row_num, column=7, value=str(route.created_by) if route.created_by else '-')

    stops_sheet = workbook.create_sheet(title='Route Stops')
    stop_headers = ['Route', 'Order', 'Bin ID', 'Bin Name', 'Location', 'Stop Status', 'Driver']
    for col_num, header in enumerate(stop_headers, 1):
        cell = stops_sheet.cell(row=1, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)

    for row_num, stop in enumerate(stops, 2):
        stops_sheet.cell(row=row_num, column=1, value=stop.route.route_name)
        stops_sheet.cell(row=row_num, column=2, value=stop.stop_order)
        stops_sheet.cell(row=row_num, column=3, value=stop.bin.bin_id)
        stops_sheet.cell(row=row_num, column=4, value=stop.bin.name)
        stops_sheet.cell(row=row_num, column=5, value=stop.bin.location_name or '-')
        stops_sheet.cell(row=row_num, column=6, value=stop.status)
        stops_sheet.cell(
            row=row_num,
            column=7,
            value=str(stop.route.assigned_driver) if stop.route.assigned_driver else '-'
        )

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="waste_management_report.xlsx"'

    workbook.save(response)
    return response


@login_required
def printable_reports_view(request):
    if request.user.role not in ['ADMIN', 'MANAGER']:
        return redirect('role_redirect')

    context = get_filtered_report_data(request)
    return render(request, 'routes/reports_print.html', context)

@login_required
def send_manual_report_view(request, report_type):
    if request.user.role not in ['ADMIN', 'MANAGER']:
        return redirect('role_redirect')

    report_data = get_report_summary(report_type=report_type)
    summary_text = build_summary_text(report_data)

    recipient_list = getattr(settings, 'MANAGER_REPORT_EMAILS', [])
    if recipient_list:
        send_mail(
            subject=f"{report_type.title()} Waste Management Report",
            message=summary_text,
            from_email=settings.DEFAULT_FROM_EMAIL,
            recipient_list=recipient_list,
            fail_silently=False,
        )

    send_telegram_message(f"<pre>{summary_text}</pre>")
    messages.success(request, f"{report_type.title()} report sent successfully.")
    return redirect('reports_home')

@csrf_exempt
def update_driver_location(request):
    if request.method != 'POST':
        return JsonResponse({'error': 'POST required'}, status=405)
    data = json.loads(request.body)
    driver_id = data.get('driver_id')
    lat = data.get('latitude')
    lng = data.get('longitude')
    if not all([driver_id, lat, lng]):
        return JsonResponse({'error': 'Missing fields'}, status=400)
    try:
        driver = CustomUser.objects.get(id=driver_id, role='DRIVER')
    except CustomUser.DoesNotExist:
        return JsonResponse({'error': 'Invalid driver'}, status=404)
    DriverLocation.objects.update_or_create(
        driver=driver,
        defaults={
            'latitude': lat,
            'longitude': lng
        }
    )
    return JsonResponse({'message': 'Location updated'})

def get_driver_location(request, driver_id):
    try:
        location = DriverLocation.objects.get(driver_id=driver_id)

        return JsonResponse({
            'latitude': location.latitude,
            'longitude': location.longitude,
            'timestamp': location.timestamp
        })
    except DriverLocation.DoesNotExist:
        return JsonResponse({'error': 'No location found'}, status=404)
    
@login_required
def live_map_view(request, route_id):
    route = Route.objects.prefetch_related('stops__bin').get(id=route_id)

    return render(request, 'routes/live_map.html', {
        'route': route
    })